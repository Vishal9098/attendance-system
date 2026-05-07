from fastapi import APIRouter, HTTPException, Depends
from fastapi.responses import StreamingResponse
from database import get_db
from auth_utils import get_current_user, require_role
from datetime import datetime, date
import io

router = APIRouter()

async def get_report_data(db, user_id: str = None, start: str = None, end: str = None, role: str = "employee"):
    query = {}
    if user_id:
        query["user_id"] = user_id
    if start:
        query["date"] = {"$gte": start}
    if end:
        query.setdefault("date", {})["$lte"] = end

    records = await db.attendance.find(
        query, {"selfie_in": 0, "selfie_out": 0, "face_encoding": 0}
    ).sort("date", -1).to_list(1000)
    for r in records:
        r["id"] = str(r.pop("_id"))
        if r.get("punch_in"):
            r["punch_in"] = r["punch_in"].strftime("%Y-%m-%d %H:%M") if isinstance(r["punch_in"], datetime) else r["punch_in"]
        if r.get("punch_out"):
            r["punch_out"] = r["punch_out"].strftime("%Y-%m-%d %H:%M") if isinstance(r["punch_out"], datetime) else r["punch_out"]
    return records

@router.get("/my-report")
async def my_report(start_date: str = None, end_date: str = None, current_user=Depends(get_current_user)):
    db = get_db()
    data = await get_report_data(db, str(current_user["_id"]), start_date, end_date)
    return {"employee": current_user["name"], "records": data, "total": len(data)}

@router.get("/team-report")
async def team_report(
    start_date: str = None,
    end_date: str = None,
    current_user=Depends(require_role("manager", "admin"))
):
    db = get_db()
    if current_user["role"] == "manager":
        team = await db.users.find({"manager_id": str(current_user["_id"])}, {"_id": 1}).to_list(200)
        user_ids = [str(u["_id"]) for u in team]
        query = {"user_id": {"$in": user_ids}}
    else:
        query = {}

    if start_date:
        query["date"] = {"$gte": start_date}
    if end_date:
        query.setdefault("date", {})["$lte"] = end_date

    records = await db.attendance.find(query, {"selfie_in": 0, "selfie_out": 0}).sort("date", -1).to_list(5000)
    for r in records:
        r["id"] = str(r.pop("_id"))
        for field in ["punch_in", "punch_out"]:
            if r.get(field) and isinstance(r[field], datetime):
                r[field] = r[field].strftime("%Y-%m-%d %H:%M")
    return {"records": records, "total": len(records)}

@router.get("/export/excel")
async def export_excel(
    start_date: str = None,
    end_date: str = None,
    current_user=Depends(get_current_user)
):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(500, "openpyxl not installed")

    db = get_db()
    uid = str(current_user["_id"]) if current_user["role"] == "employee" else None
    records = await get_report_data(db, uid, start_date, end_date)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    headers = ["Date", "Employee", "Punch In", "Punch Out", "Hours Worked", "Status"]
    header_fill = PatternFill("solid", fgColor="2563EB")
    header_font = Font(bold=True, color="FFFFFF")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = 18

    for row, r in enumerate(records, 2):
        ws.cell(row, 1, r.get("date", ""))
        ws.cell(row, 2, r.get("user_name", ""))
        ws.cell(row, 3, r.get("punch_in", ""))
        ws.cell(row, 4, r.get("punch_out", ""))
        ws.cell(row, 5, r.get("hours_worked", ""))
        ws.cell(row, 6, r.get("status", ""))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=attendance_report.xlsx"}
    )

@router.get("/export/pdf")
async def export_pdf(
    start_date: str = None,
    end_date: str = None,
    current_user=Depends(get_current_user)
):
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib import colors
    except ImportError:
        raise HTTPException(500, "reportlab not installed")

    db = get_db()
    uid = str(current_user["_id"]) if current_user["role"] == "employee" else None
    records = await get_report_data(db, uid, start_date, end_date)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4))
    styles = getSampleStyleSheet()
    elements = [Paragraph("Attendance Report", styles["Title"])]

    table_data = [["Date", "Employee", "Punch In", "Punch Out", "Hours", "Status"]]
    for r in records:
        table_data.append([
            r.get("date", ""),
            r.get("user_name", ""),
            r.get("punch_in", ""),
            r.get("punch_out", ""),
            str(r.get("hours_worked", "")),
            r.get("status", "")
        ])

    t = Table(table_data, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#2563EB")),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#F1F5F9")]),
        ("GRID", (0,0), (-1,-1), 0.5, colors.grey),
        ("ALIGN", (0,0), (-1,-1), "CENTER"),
    ]))
    elements.append(t)
    doc.build(elements)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=attendance_report.pdf"}
    )
